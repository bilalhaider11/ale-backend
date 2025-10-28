import os
import tempfile
import csv
from flask import request
from flask_restx import Namespace, Resource
from openpyxl import load_workbook
from common.models.alert import AlertLevelEnum, AlertStatusEnum
from common.app_config import config
from common.app_logger import logger
from common.services.employee import EmployeeService
from common.services.organization import OrganizationService
from common.services.alert import AlertService
from app.helpers.response import get_success_response, get_failure_response, parse_request_body, validate_required_fields
from app.helpers.decorators import login_required, organization_required, with_partner_organization_ids
from common.models.person_organization_role import PersonOrganizationRoleEnum
from common.models import Employee
from common.services import (
    PersonOrganizationInvitationService,
    PersonService,
    OrganizationService,
    AvailabilitySlotService
)
from common.models import Person
from datetime import datetime

employee_api = Namespace('employee', description='Employee operations')


@employee_api.route('/upload')
class EmployeeListUpload(Resource):

    @login_required()
    @organization_required(with_roles=[PersonOrganizationRoleEnum.ADMIN])
    def post(self, person, organization):
        """
        Upload a CSV or XLSX file with employee/caregiver or physician data.
        If file contains NPI column, it will be processed as physician data.
        Otherwise, it will be processed as employee data.
        """
        employee_service = EmployeeService(config)
        organization_service = OrganizationService(config)
        alert_service = AlertService(config)

        if 'file' not in request.files:
            return get_failure_response("No file provided", status_code=400)
        
        file = request.files['file']
        file_id = request.form.get('file_id', None)
        
        if not file.filename:
            return get_failure_response("No file selected", status_code=400)
        
        # Check allowed extensions
        allowed_extensions = ['.csv', '.xlsx']
        if not any(file.filename.lower().endswith(ext) for ext in allowed_extensions):
            return get_failure_response("File must be a CSV or XLSX", status_code=400)
        
        # Save file temporarily with appropriate extension
        file_extension = '.csv' if file.filename.lower().endswith('.csv') else '.xlsx'
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as temp_file:
            file.save(temp_file.name)
            temp_file_path = temp_file.name
        
        try:
            #Get all existing employee IDs
            employee_ids = employee_service.get_all_employee_ids(organization.entity_id)
            print(employee_ids)
            existing_ids = {str(row['employee_id']) for row in employee_ids}
            print("Existing IDs:", existing_ids)

            file_category = "employee"

            if file_extension == '.xlsx':
                workbook = load_workbook(temp_file_path, data_only=False)
                worksheet = workbook.active

                for row_idx, row in enumerate(worksheet.iter_rows(values_only=True,min_row=2), start=2):
                    print("///////////////////////////////////////////")
                    print(row)

                    row = list(row)

                    # for physician data
                    if row and any(cell and 'npi' in str(cell).lower() for cell in row):
                        file_category = "physician"
                        break

                    #if not any(row):
                    #    continue

                    # Assign new employee_id to missing employee ids
                    if row[0] is None:
                        print("/////////////////////////////////////////////////////")
                        print("/////////////////////////////////////////////////////")
                        print("/////////////////////////////////////////////////////")
                        print("/////////////////////////////////////////////////////")
                        new_emp_id = organization_service.get_next_employee_id(organization.entity_id)
                        row[0] = new_emp_id
                        existing_ids.add(new_emp_id)
                        print("existing_ids: ",existing_ids)
                        
                        workbook = load_workbook(temp_file_path)
                        sheet = workbook.active
                        
                        cell = sheet.cell(row=row_idx, column=1, value=new_emp_id)
                        
                        workbook.save(temp_file_path)
                        
                        
                        print(f"Generated new employee_id {new_emp_id} for row {row_idx}")
                    else:
                        current_emp_id = str(row[0])
                        print(f"Found employee_id: {current_emp_id}")
                        

                        # Check for duplicate
                        
                        if current_emp_id in existing_ids:
                            
                            logger.warning(f"Duplicate employee_id detected: {current_emp_id}")

                            description = (
                                f"Duplicate employee ID detected: {current_emp_id} for organization {organization.entity_id}."
                            )
                            status_ = AlertStatusEnum.ADDRESSED.value
                            level = AlertLevelEnum.WARNING.value
                            title = "Employee"

                            
                            alert_service.create_alert(
                                organization_id=organization.entity_id,
                                title=title,
                                description=description,
                                alert_type=level,
                                status=status_,
                                assigned_to_id=current_emp_id
                            )
                            print(f" Created alert for duplicate employee_id: {current_emp_id}")
                            
                        else:
                            existing_ids.add(current_emp_id)
                            print("existing_ids: ",existing_ids)
                                
                    print(",,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,")
                    print(row)
                            
                    

            else:
                # Handle CSV files similarly
                with open(temp_file_path, 'r', encoding='utf-8') as csv_file:
                    reader = csv.reader(csv_file)
                    for i, row in enumerate(reader):
                        # optional: skip header detection or use a smarter header check
                        if i == 0 and any(cell and 'employee_id' in cell.lower() for cell in row):
                            continue

                        if row and any(cell and 'npi' in cell.lower() for cell in row):
                            file_category = "physician"
                            break

                        if not row:
                            continue

                        if not row[0]:
                            new_emp_id = organization_service.get_next_employee_id(organization.entity_id)
                            row[0] = new_emp_id
                        else:
                            current_emp_id = str(row[0])
                            if current_emp_id in existing_ids:
                                logger.warning(f"Duplicate employee_id detected: {current_emp_id}")
                                description = (
                                    f"Duplicate employee ID detected: {current_emp_id} for organization {organization.entity_id}."
                                )
                                status_ = AlertStatusEnum.ADDRESSED.value
                                level = AlertLevelEnum.WARNING.value
                                title = "Employee"

                                try:
                                    alert_service.create_alert(
                                        organization_id=organization.entity_id,
                                        title=title,
                                        description=description,
                                        alert_type=level,
                                        status=status_,
                                        assigned_to_id=current_emp_id
                                    )
                                except Exception as alert_err:
                                    logger.exception(f"Failed to create duplicate alert: {alert_err}")


                        
            workbook = load_workbook(filename=temp_file_path) 
            sheet = workbook.active 
                
            for row in sheet.iter_rows(values_only=True):
                print(row)


            # Step 3: Upload to S3 after processing
            upload_result = employee_service.upload_list_file(
                organization_id=organization.entity_id,
                person_id=person.entity_id,
                file_path=temp_file_path,
                file_category=file_category,
                file_id=file_id,
                original_filename=file.filename
            )

            os.unlink(temp_file_path)
            
            return get_success_response(
                message=f"File uploaded successfully as {file_category} data",
                upload_info=upload_result,
                file_category=file_category
            )

        except Exception as e:
            if os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
            logger.exception(e)
            return get_failure_response(f"Error uploading file: {str(e)}", status_code=500)

@employee_api.route('/matches')
class EmployeeListMatches(Resource):

    @login_required()
    @organization_required(with_roles=[PersonOrganizationRoleEnum.ADMIN])
    def get(self, person, organization):
        """
        Get a list of employees who have matches in the employee_exclusion_match table.
        """
        employee_service = EmployeeService(config)
        matched_employees = employee_service.get_employees_with_matches(organization.entity_id)

        return get_success_response(
            message="Matched employees retrieved successfully",
            data=matched_employees
        )

@employee_api.route('/admin')
class EmployeeAdmin(Resource):

    @login_required()
    @organization_required(with_roles=[PersonOrganizationRoleEnum.ADMIN])
    @with_partner_organization_ids()
    def get(self, person, organization, partner_organization_ids):
        # Get employee_type from query parameters
        employee_type = request.args.get('employee_type', None)
        
        employee_service = EmployeeService(config)
        employees = employee_service.get_employees_by_organization(partner_organization_ids, employee_type=employee_type)
        return get_success_response(
            message="Employees retrieved successfully",
            data=employees
        )

@employee_api.route('/admin/<string:entity_id>')
class EmployeeDetails(Resource):
    @login_required()
    @organization_required(with_roles=[PersonOrganizationRoleEnum.ADMIN])
    def get(self, person, organization, entity_id):
        """
        Get a single employee by their entity_id.
        """
        employee_service = EmployeeService(config)
        employee = employee_service.get_employee_by_id(entity_id, organization.entity_id)

        if not employee:
            return get_failure_response("Employee not found", status_code=404)

        return get_success_response(
            message="Employee retrieved successfully",
            data=employee
        )

@employee_api.route('')
class EmployeeResource(Resource):

    @login_required()
    @organization_required(with_roles=[PersonOrganizationRoleEnum.EMPLOYEE])
    def get(self, person, organization):

        employee_service = EmployeeService(config)
        employee = employee_service.get_employee_by_person_id(person.entity_id, organization.entity_id)
        return get_success_response(
            message="Employee rerieved successfully",
            data=employee
        )

    @login_required()
    @organization_required(with_roles=[PersonOrganizationRoleEnum.ADMIN])
    def post(self, person, organization):
        """
        Create a new employee record.
        """
        employee_service = EmployeeService(config)
        person_service = PersonService(config)
        organization_service = OrganizationService(config)
        
        # Parse request body
        parsed_body = parse_request_body(request, [
            'first_name',
            'last_name',
            'employee_id',
            'date_of_birth',
            'email_address',
            'phone_1',
            'employee_type'
        ])

        date_of_birth = parsed_body.pop('date_of_birth', None)
        email_address = parsed_body.pop('email_address', None)
        phone_1 = parsed_body.pop('phone_1', None)
        employee_id = parsed_body.pop('employee_id', None)
        
        validate_required_fields(parsed_body)
        
        # Auto-generate employee_id if not provided
        if not employee_id or not employee_id.strip():
            employee_id = organization_service.get_next_employee_id(organization.entity_id)
        
        # Check for duplicate employee ID
        existing_employee = employee_service.employee_repo.get_by_employee_id(employee_id, organization.entity_id)
        
        print("existing_employee: ",existing_employee)
        if existing_employee:
            
            description = f"Duplicate employee ID detected: Duplicate employee ID detected: ${employee_id} for organization ${organization.entity_id}.${employee_id} for organization ${organization.entity_id}."
            status_ =  AlertStatusEnum.ADDRESSED.value
            level = AlertLevelEnum.WARNING.value
            title = 'Employee'
            logger.warning(
                f"Duplicate employee ID detected: {employee_id} for organization {organization.entity_id}. "
                f"Existing employee: {existing_employee.entity_id} ({existing_employee.first_name} {existing_employee.last_name}). "
                f"Creating new employee anyway as per requirements."
                
            )
            
            alert_service = AlertService(config)
            try:
                print(alert_service)
                create_duplicateRecord_alert = alert_service.create_alert(
                    organization_id = organization.entity_id,
                    title = title,
                    description = description,
                    alert_type = level,
                    status = status_,
                    assigned_to_id = employee_id
                
                )
            except Exception as e:
                logger.error(f"Error processing patient file: {str(e)}")
            
        
        person = Person(
            first_name=parsed_body['first_name'],
            last_name=parsed_body['last_name']
        )
        person = person_service.save_person(person)

        employee = Employee(
            first_name=parsed_body['first_name'],
            last_name=parsed_body['last_name'],
            employee_id=employee_id,
            date_of_birth=date_of_birth,
            email_address=email_address,
            phone_1=phone_1,
            organization_id=organization.entity_id,
            person_id=person.entity_id,
            employee_type=parsed_body['employee_type']
        )
        employee = employee_service.save_employee(employee)
        
        employee_service.trigger_match_for_employee(employee.entity_id)
        return get_success_response(
            message="Employee created successfully",
            data=employee.as_dict()
        )

    @login_required()
    @organization_required(with_roles=[PersonOrganizationRoleEnum.ADMIN])
    def put(self, person, organization):
        """
        Update an existing employee record.
        """
        employee_service = EmployeeService(config)
        person_service = PersonService(config)
        organization_service = OrganizationService(config)
        
        # Parse request body
        parsed_body = parse_request_body(request, [
            'entity_id',
            'first_name',
            'last_name',
            'employee_id',
            'date_of_birth',
            'email_address',
            'phone_1',
            'employee_type'
        ])

        entity_id = parsed_body.pop('entity_id', None)
        date_of_birth = parsed_body.pop('date_of_birth', None)
        email_address = parsed_body.pop('email_address', None)
        phone_1 = parsed_body.pop('phone_1', None)
        employee_id = parsed_body.pop('employee_id', None)
        
        if not entity_id:
            return get_failure_response("entity_id is required for update", status_code=400)
        
        validate_required_fields(parsed_body)
        
        # Get existing employee
        employee = employee_service.get_employee_by_id(entity_id, organization.entity_id)
        if not employee:
            return get_failure_response("Employee not found", status_code=404)
        
        employee.first_name = parsed_body['first_name']
        employee.last_name = parsed_body['last_name']
        
        # Auto-generate employee_id if not provided
        if not employee_id or not employee_id.strip():
            employee_id = organization_service.get_next_employee_id(organization.entity_id)
        
        employee.employee_id = employee_id
        employee.date_of_birth = date_of_birth
        employee.email_address = email_address
        employee.phone_1 = phone_1
        employee.employee_type = parsed_body['employee_type']

        if employee.person_id:
            person = person_service.get_person_by_id(employee.person_id)
            if person and (person.first_name != employee.first_name or person.last_name != employee.last_name):
                person.first_name = employee.first_name
                person.last_name = employee.last_name
                person_service.save_person(person)
        else:
            person = Person(
                first_name=employee.first_name,
                last_name=employee.last_name,
            )
            person = person_service.save_person(person)
            employee.person_id = person.entity_id

        employee = employee_service.save_employee(employee)
        
        employee_service.trigger_match_for_employee(employee.entity_id)
        return get_success_response(
            message="Employee updated successfully",
            data=employee.as_dict()
        )


@employee_api.route('/invite')
class EmployeeInvite(Resource):

    @login_required()
    @organization_required(with_roles=[PersonOrganizationRoleEnum.ADMIN])
    def post(self, person, organization):
        """
        Invite an employee to join the organization.
        """
        invitation_service = PersonOrganizationInvitationService(config)
        person_service = PersonService(config)
        employee_service = EmployeeService(config)
        organization_service = OrganizationService(config)

        # Parse request body
        parsed_body = parse_request_body(request, ['employee_id'])
        validate_required_fields({'employee_id': parsed_body['employee_id']})
        employee_id = parsed_body['employee_id']

        # Get the employee record
        employee = employee_service.get_employee_by_id(employee_id, organization.entity_id)
        if not employee:
            return get_failure_response(message="Employee not found.", status_code=404)
        
        if not employee.email_address:
            return get_failure_response(message="Employee must have an email address to send invitation.", status_code=200)

        # Get the organization record
        org = organization_service.get_organization_by_id(organization.entity_id)
        if not org:
            return get_failure_response(message="Organization not found.", status_code=404)
        
        employee_person = person_service.get_person_by_id(employee.person_id)

        existing_person = person_service.get_person_by_email_address(employee.email_address)
        if existing_person and existing_person.entity_id != employee_person.entity_id:
            if existing_person.first_name.strip().lower() != employee_person.first_name.strip().lower() or \
                existing_person.last_name.strip().lower() != employee_person.last_name.strip().lower():
                # If names don't match, we have a potential conflict
                return get_failure_response(message="Email address is already in use by another person under a different name.", status_code=200)

            invited_person = existing_person
        else:
            invited_person = employee_person

        # Create and send invitation with employee role
        invitation = invitation_service.create_invitation(
            organization_id=organization.entity_id,
            invitee_id=invited_person.entity_id,
            email=employee.email_address,
            roles=['employee'],
            first_name=employee.first_name,
            last_name=employee.last_name,
            invited_by_id=person.entity_id
        )
        invitation_service.send_invitation_email(invitation, org.name, person)

        employee.person_id = invited_person.entity_id
        employee_service.save_employee(employee)

        return get_success_response(message="Employee invitation sent successfully.")



@employee_api.route('/by/slot')
class EmployeesBySlot(Resource):

    @login_required()
    @organization_required(with_roles=[PersonOrganizationRoleEnum.ADMIN])
    def get(self, person, organization):
        """
        Get all employees for the organization.
        """

        date = request.args['date']
        slot_start_time = request.args['start_time']
        slot_end_time = request.args['end_time']
        patient_id = request.args['patient_id']

        # Parse date and time strings
        try:
            date = datetime.strptime(date, '%Y-%m-%d').date()
            slot_start_time = datetime.strptime(slot_start_time, '%H:%M').time()
            slot_end_time = datetime.strptime(slot_end_time, '%H:%M').time()
        except ValueError:
            return get_failure_response("Invalid date or time format", status_code=400)

        availability_slot_service = AvailabilitySlotService(config)
        availability_slots = availability_slot_service.get_availability_slots_for_time_slot(
            start_time=slot_start_time,
            end_time=slot_end_time,
            visit_date=date,
            patient_id=patient_id,
            organization_ids=[organization.entity_id]
        )

        return get_success_response(
            slots=availability_slots,
            count=len(availability_slots)
        )
