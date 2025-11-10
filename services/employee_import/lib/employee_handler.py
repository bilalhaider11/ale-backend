import tempfile
import os
import csv
from openpyxl import load_workbook
from common.app_logger import create_logger
from common.services.employee import EmployeeService
from common.services.current_employees_file import CurrentEmployeesFileService
from common.services.s3_client import S3ClientService
from common.models.current_employees_file import CurrentEmployeesFileStatusEnum
logger = create_logger()

class EmployeeHandler:
    """Handler for processing employee CSV files"""
    
    def __init__(self, config):
        self.config = config
        self.s3_client = S3ClientService()
        self.employee_service = EmployeeService(config)
        self.employees_file_service = CurrentEmployeesFileService(config)

    def _read_excel_to_dict_list(self, file_path):
        """
        Read Excel file and convert to list of dictionaries
        
        Args:
            file_path (str): Path to the Excel file
            
        Returns:
            list: List of dictionaries representing rows
        """
        # Only first name and last name are required now
        # Employee ID is optional and will be auto-generated if missing
        required_headers = [
            'first_name', 
            'last_name'
        ]
        
        workbook = load_workbook(file_path, data_only=True)
        worksheet = workbook.active
        
        header_row_index = None
        header_row = None
        
        # Loop through rows to find the header row
        for row_num, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue  # Skip empty rows
            
            # Convert row headers to lowercase for case-insensitive comparison
            row_headers_lower = [str(cell).lower().strip() if cell is not None else '' for cell in row]
            
            # Check if all required headers are present
            has_all_required = True
            
            for required_header in required_headers:
                header_opts = required_header
                if isinstance(header_opts, str):
                    header_opts = (required_header, )

                match_found = False
                for opt in header_opts:
                    for row_header in row_headers_lower:
                        if opt in row_header:
                            match_found = True
                            break
                    if match_found:
                        break
                if not match_found:
                    has_all_required = False
                    break

            if has_all_required:
                header_row_index = row_num
                header_row = [str(cell).lower().strip() if cell is not None else '' for cell in row]
                break
        
        if header_row_index is None:
            raise ValueError("first name and last name headers are required")
        
        # Convert rows to dictionaries starting from the header row onwards
        rows = []
        for row in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue  # Skip empty rows
            
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(header_row) and header_row[i]:
                    row_dict[header_row[i]] = value
            rows.append(row_dict)

        if not rows:
            raise ValueError("No data found in the file.")

        return rows
    
    def _read_csv_to_dict_list(self, file_path: str) -> list[dict]:
        """
        Read CSV file and convert to list of dicts.
        Required headers with multiple variations for each field.
        """
        required_headers = [
            ("first_name", "first name", "f name", "f_name"),
            ("last_name", "last name", "l name", "l_name"),
            ("date of birth", "date_of_birth", "dob"),
            ("employee id", "employee_id", "id"),
            ("phone","mobile","contact","contact no","contact_no","mobile no","mobile_no","cell","cell no","cell_no"),
            ("email","e-mail","email_address","email address")
            
        ]
    
        with open(file_path, mode="r", encoding="utf-8") as f:
            reader = csv.reader(f)
            all_rows = list(reader)
    
        header_row_index = None
        header_row = None
    
        for i, row in enumerate(all_rows):
            if not row:
                continue
                
            row_headers_lower = [cell.lower().strip() if cell else "" for cell in row]
            
            # Check if all required header groups have at least one match
            has_all_required = True
            for header_group in required_headers:
                if not any(variant in row_headers_lower for variant in header_group):
                    has_all_required = False
                    break
                    
            if has_all_required:
                header_row_index = i
                header_row = row_headers_lower
                break
    
        if header_row_index is None:
            raise ValueError("Error in headers name of your file, download the sample file provided ,to check the headers name or use these instead: ",required_headers)
    
        # Process data rows
        data_rows = all_rows[header_row_index + 1:]
        result = []
        
        for row in data_rows:
            if not row or all(not (cell or "").strip() for cell in row):
                continue
                
            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(header_row) and header_row[i]:
                    # Normalize header names to standard format
                    header_value = header_row[i]
                    for header_group in required_headers:
                        if header_value in header_group:
                            # Use the first variant as the standardized key
                            row_dict[header_group[0]] = cell
                            break
                    else:
                        # If not in required headers, use as-is
                        row_dict[header_value] = cell
                        
            result.append(row_dict)
    
        if not result:
            raise ValueError("No Data exist in your .csv file")
    
        return result

    def process_employee_list(self, key, file_category):
        """
        Process an employee or physician CSV or XLSX file from S3
        
        Args:
            key (str): S3 object key
            file_category (str): Either "employee" or "physician"
            
        Returns:
            bool: True if successful, False otherwise
        """
        # Determine file content type from metadata and create appropriate temporary file
        _, organization_id, file_id = key.rsplit('/', 2)

        employees_file = self.employees_file_service.get_by_id(file_id, organization_id)
        file_extension = f".{employees_file.file_type.lower()}"

        object_metadata = self.s3_client.get_object_metadata(key)

        organization_id = object_metadata.get('organization_id')

        self.employees_file_service.update_status(employees_file, CurrentEmployeesFileStatusEnum.PROCESSING)

        with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as temp_file:
            temp_path = temp_file.name

        self.s3_client.download_file(key, temp_path)
        
        # Read file using appropriate method based on extension
        try:
            try:
                if file_extension == '.xlsx':
                    # Read XLSX file using openpyxl
                    rows = self._read_excel_to_dict_list(temp_path)
                    
                else:
                    # Read CSV file using csv module
                    rows = self._read_csv_to_dict_list(temp_path)
            except ValueError as e:
                self.employees_file_service.set_error(
                    employees_file, str(e)
                )
                return False
            
            logger.info(f"Found {len(rows)} {file_category} records in file")
            
            # Import new data
            if file_category == "physician":
                from common.services.physician import PhysicianService
                physician_service = PhysicianService(self.config)
                import_count, skipped_entries = physician_service.bulk_import_physicians(rows, organization_id=organization_id, user_id=employees_file.uploaded_by)
            else:
                import_count, skipped_entries = self.employee_service.bulk_import_employees(rows, organization_id=organization_id, user_id=employees_file.uploaded_by)
            employees_file.record_count = import_count
            self.employees_file_service.update_status(employees_file, CurrentEmployeesFileStatusEnum.IMPORTED)
            return True

        finally:
            # Clean up the temporary file
            if os.path.exists(temp_path):
                os.unlink(temp_path)
