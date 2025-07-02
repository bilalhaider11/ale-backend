import tempfile
import os
import csv
from openpyxl import load_workbook
from common.app_logger import create_logger
from common.services.current_employee import CurrentEmployeeService
from common.services.current_employees_file import CurrentEmployeesFileService
from common.services.s3_client import S3ClientService
from common.models.current_employees_file import CurrentEmployeesFileStatusEnum

logger = create_logger()

class CurrentEmployeeHandler:
    """Handler for processing employee CSV files"""
    
    def __init__(self, config):
        self.config = config
        self.s3_client = S3ClientService()
        self.employee_service = CurrentEmployeeService(config)
        self.employees_file_service = CurrentEmployeesFileService(config)

    def _read_excel_to_dict_list(self, file_path):
        """
        Read Excel file and convert to list of dictionaries
        
        Args:
            file_path (str): Path to the Excel file
            
        Returns:
            list: List of dictionaries representing rows
        """
        required_headers = ['first name', 'last name', 'date of birth']
        
        workbook = load_workbook(file_path, data_only=True)
        worksheet = workbook.active
        
        header_row_index = None
        header_row = None
        
        # Loop through rows to find the header row
        for row_num, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue  # Skip empty rows
            
            # Convert row headers to lowercase for case-insensitive comparison
            row_headers_lower = [str(cell).lower().strip() if cell is not None else '' for cell in row]
            
            # Check if all required headers are present
            has_all_required = all(
                any(required_header in row_header for row_header in row_headers_lower)
                for required_header in required_headers
            )
            
            if has_all_required:
                header_row_index = row_num
                header_row = [str(cell).lower().strip() if cell is not None else '' for cell in row]
                break
        
        if header_row_index is None:
            raise ValueError("ENOHEADERS")
        
        # Convert rows to dictionaries starting from the header row onwards
        rows = []
        for row in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue  # Skip empty rows
            
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(header_row) and header_row[i]:
                    row_dict[header_row[i]] = value
            rows.append(row_dict)

        if not rows:
            raise ValueError("ENODATA")

        return rows
    

    def _read_csv_to_dict_list(self, file_path):
        """
        Read CSV file and convert to list of dictionaries
        
        Args:
            file_path (str): Path to the CSV file
            
        Returns:
            list[dict]: List of dictionaries representing rows
        """

        required_headers = ['first name', 'last name', 'date of birth']
        
        with open(file_path, mode='r', encoding='utf-8') as csv_file:
            reader = csv.reader(csv_file)
            all_rows = list(reader)
            
            header_row_index = None
            header_row = None
            
            # Find the header row
            for i, row in enumerate(all_rows):
                if not row:
                    continue
                
                row_headers_lower = [cell.lower().strip() if cell else '' for cell in row]
                if all(req in row_headers_lower for req in required_headers):
                    header_row_index = i
                    header_row = [cell.strip().lower() if cell else '' for cell in row]  # Normalize
                    break
            
            if header_row_index is None:
                raise ValueError("ENOHEADERS")
            
            data_rows = all_rows[header_row_index + 1:]
            
            result = []
            for row in data_rows:
                if not row or all(not (cell or "").strip() for cell in row):
                    continue
                row_dict = {
                    header_row[i]: cell
                    for i, cell in enumerate(row)
                    if i < len(header_row) and header_row[i]
                }
                result.append(row_dict)
            
            if not result:
                raise ValueError("ENODATA")

            return result


    def process_employee_list(self, key):
        """
        Process an employee CSV or XLSX file from S3
        
        Args:
            bucket (str): S3 bucket name
            key (str): S3 object key
            
        Returns:
            bool: True if successful, False otherwise
        """
        # Determine file content type from metadata and create appropriate temporary file
        _, organization_id, file_id = key.rsplit('/', 2)

        employees_file = self.employees_file_service.get_by_id(file_id, organization_id)
        file_extension = f".{employees_file.file_type.lower()}"

        object_metadata = self.s3_client.get_object_metadata(key)

        organization_id = object_metadata.get('organization_id')

        self.employees_file_service.update_status(employees_file, CurrentEmployeesFileStatusEnum.PROCESSING)

        with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as temp_file:
            temp_path = temp_file.name

        self.s3_client.download_file(key, temp_path)
        
        # Read file using appropriate method based on extension
        try:
            try:
                if file_extension == '.xlsx':
                    # Read XLSX file using openpyxl
                    rows = self._read_excel_to_dict_list(temp_path)
                else:
                    # Read CSV file using csv module
                    rows = self._read_csv_to_dict_list(temp_path)
            except ValueError as e:
                self.employees_file_service.set_error(
                    employees_file, str(e)
                )
                return False
            
            logger.info(f"Found {len(rows)} employee records in file")
            
            # Import new data
            import_count = self.employee_service.bulk_import_employees(rows, organization_id=organization_id)
            employees_file.record_count = import_count
            self.employees_file_service.update_status(employees_file, CurrentEmployeesFileStatusEnum.IMPORTED)

            return True

        finally:
            # Clean up the temporary file
            if os.path.exists(temp_path):
                os.unlink(temp_path)
