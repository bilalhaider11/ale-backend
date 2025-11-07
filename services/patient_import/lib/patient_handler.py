import tempfile
import os
import csv
from openpyxl import load_workbook

from common.app_logger import create_logger
from common.services.patient import PatientService
from common.services.patients_file import PatientsFileService
from common.services.s3_client import S3ClientService
from common.models.patients_file import PatientsFileStatusEnum

logger = create_logger()


class PatientHandler:
    """Handler for processing patient CSV/XLSX files"""

    def __init__(self, config):
        self.config = config
        self.s3_client = S3ClientService()
        self.patient_service = PatientService(config)
        self.patient_file_service = PatientsFileService(config)

    def _read_excel_to_dict_list(self, file_path: str) -> list[dict]:
        """
        Read Excel file and convert to list of dicts.
        Required headers: first name, last name, date of birth (any variant).
        """
        required_headers = [
            ("first_name", "first name"),
            ("last_name", "last name"),
            ("date of birth", "date_of_birth", "dob"),
            ("medical record number", "medical_record_number", "mrn"),
            ("gender","sex")
        ]

        workbook = load_workbook(file_path, data_only=True)
        worksheet = workbook.active

        header_row_index = None
        header_row = None

        for row_num, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            row_headers_lower = [str(cell).lower().strip() if cell is not None else "" for cell in row]

            has_all_required = True
            for required in required_headers:
                opts = (required,) if isinstance(required, str) else required
                if not any(opt in h for opt in opts for h in row_headers_lower):
                    has_all_required = False
                    break

            if has_all_required:
                header_row_index = row_num
                header_row = row_headers_lower
                break

        if header_row_index is None:
            raise ValueError("ENOHEADERS")

        rows = []
        for row in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(header_row) and header_row[i]:
                    row_dict[header_row[i]] = value
            rows.append(row_dict)

        if not rows:
            raise ValueError("ENODATA")

        return rows

    def _read_csv_to_dict_list(self, file_path: str) -> list[dict]:
        """
        Read CSV file and convert to list of dicts.
        Required headers: first_name, last_name, date_of_birth (any variant).
        """
        required_headers = ["first_name", "last_name", "date_of_birth"]

        with open(file_path, mode="r", encoding="utf-8") as f:
            reader = csv.reader(f)
            all_rows = list(reader)

        header_row_index = None
        header_row = None

        for i, row in enumerate(all_rows):
            if not row:
                continue
            row_headers_lower = [cell.lower().strip() if cell else "" for cell in row]
            if all(req in row_headers_lower for req in required_headers):
                header_row_index = i
                header_row = row_headers_lower
                break

        if header_row_index is None:
            raise ValueError("ENOHEADERS") 

        data_rows = all_rows[header_row_index + 1 :]
        result = []
        for row in data_rows:
            if not row or all(not (cell or "").strip() for cell in row):
                continue
            row_dict = {
                header_row[i]: cell
                for i, cell in enumerate(row)
                if i < len(header_row) and header_row[i]
            }
            result.append(row_dict)

        if not result:
            raise ValueError("ENODATA")

        return result

    def process_patient_list(self, key: str) -> bool:
        """
        Process a patient CSV or XLSX file from S3.
        """

        _, organization_id, file_id = key.rsplit("/", 2)

        logger.info(f"File and Organization ID {file_id}, {organization_id}")

        patient_file = self.patient_file_service.get_by_id(file_id, organization_id)

        logger.info(f"File Object {patient_file}")


        file_extension = f".{patient_file.file_type.lower()}"

        # metadata can override org_id if needed
        object_metadata = self.s3_client.get_object_metadata(key)
        organization_id = object_metadata.get("organization_id") or organization_id

        self.patient_file_service.update_status(patient_file, PatientsFileStatusEnum.PROCESSING)

        # download to temp
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as tmp:
            temp_path = tmp.name

        self.s3_client.download_file(key, temp_path)

        try:
            try:
                if file_extension == ".xlsx":
                    rows = self._read_excel_to_dict_list(temp_path)
                else:
                    rows = self._read_csv_to_dict_list(temp_path)
            except ValueError as e:
                self.patient_file_service.set_error(patient_file, str(e))
                return False

            logger.info(f"Found {len(rows)} patient records in file")

            import_count = self.patient_service.bulk_import_patients(
                rows,
                organization_id=organization_id,
                user_id=patient_file.uploaded_by,
            )

            patient_file.record_count = import_count
            self.patient_file_service.update_status(patient_file, PatientsFileStatusEnum.IMPORTED)
            return True

        except Exception as e:
            logger.exception("Error processing patient file: %s", e)
            self.patient_file_service.set_error(patient_file, str(e))
            return False

        finally:
            if os.path.exists(temp_path):
                os.unlink(temp_path)
